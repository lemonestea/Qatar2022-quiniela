import operator
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


dir_path = r"C:\Users\Cesar\Desktop\Qatar 2022\QuinielaProject"
dir_quinielas = dir_path + r"\quinielas"
count = 0
resultados = openpyxl.load_workbook(dir_path + r"\resultados.xlsx")
resultados_sheet = resultados.active
posiciones = openpyxl.Workbook()
posiciones_sheet = posiciones.active
players_list = list()


def faseDeGrupos():

    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active
        player_name = sheet_obj.cell(row=2, column=6).value

        for i in range(48):

            if (i <= 5):

                score1 = sheet_obj.cell(row=6 + i, column=3).value
                score2 = sheet_obj.cell(row=6 + i, column=4).value
                result_score1 = resultados_sheet.cell(row=6 + i, column=3).value
                result_score2 = resultados_sheet.cell(row=6 + i, column=4).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    if (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    if (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 5 and i <= 11):
                score1 = sheet_obj.cell(row=i, column=8).value
                score2 = sheet_obj.cell(row=i, column=9).value
                result_score1 = resultados_sheet.cell(row=i, column=8).value
                result_score2 = resultados_sheet.cell(row=i, column=9).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    if (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    if (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 11 and i <= 17):
                score1 = sheet_obj.cell(row=i - 6, column=13).value
                score2 = sheet_obj.cell(row=i - 6, column=14).value
                result_score1 = resultados_sheet.cell(row=i - 6, column=13).value
                result_score2 = resultados_sheet.cell(row=i - 6, column=14).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 17 and i <= 23):
                score1 = sheet_obj.cell(row=i - 12, column=18).value
                score2 = sheet_obj.cell(row=i - 12, column=19).value
                result_score1 = resultados_sheet.cell(row=i - 12, column=18).value
                result_score2 = resultados_sheet.cell(row=i - 12, column=19).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 23 and i <= 29):
                score1 = sheet_obj.cell(row=i - 8, column=3).value
                score2 = sheet_obj.cell(row=i - 8, column=4).value
                result_score1 = resultados_sheet.cell(row=i - 8, column=3).value
                result_score2 = resultados_sheet.cell(row=i - 8, column=4).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 29 and i <= 35):
                score1 = sheet_obj.cell(row=i - 14, column=8).value
                score2 = sheet_obj.cell(row=i - 14, column=9).value
                result_score1 = resultados_sheet.cell(row=i - 14, column=8).value
                result_score2 = resultados_sheet.cell(row=i - 14, column=9).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 35 and i <= 41):
                score1 = sheet_obj.cell(row=i - 20, column=13).value
                score2 = sheet_obj.cell(row=i - 20, column=14).value
                result_score1 = resultados_sheet.cell(row=i - 20, column=13).value
                result_score2 = resultados_sheet.cell(row=i - 20, column=14).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

            if (i > 41 and i <= 47):
                score1 = sheet_obj.cell(row=i - 26, column=18).value
                score2 = sheet_obj.cell(row=i - 26, column=19).value
                result_score1 = resultados_sheet.cell(row=i - 26, column=18).value
                result_score2 = resultados_sheet.cell(row=i - 26, column=19).value

                if (result_score1 != None and result_score2 != None):

                    if (score1 > score2 and result_score1 > result_score2):
                        points += 2

                    elif (score1 < score2 and result_score1 < result_score2):
                        points += 2

                    elif (score1 == score2 and result_score1 == result_score2):
                        points += 2

                    if (score1 == result_score1):
                        points += 1

                    if (score2 == result_score2):
                        points += 1

        #for i in range(16):

        player = {"player name": player_name, "points_fg": points, "points_octavos" : 0,
                  "points_cuartos": 0, "points_semis": 0,"points_tercero": 0,"points_final": 0,
                  "points_total": points}
        global players_list
        players_list.append(player)

        wb_obj.close()

def clasificadosOctavos():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active

        for i in range(8):

            if (i <= 3):

                player_first = sheet_obj.cell(row=13, column=2 + (i * 5)).value
                player_second = sheet_obj.cell(row=13, column=4 + (i * 5)).value
                result_first = resultados_sheet.cell(row=13, column=2 + (i * 5)).value
                result_second = resultados_sheet.cell(row=13, column=4 + (i * 5)).value

                if (result_first != None and result_second != None):

                    if (player_first.lower() == result_first.lower()):
                        points += 3

                    if (player_second.lower() == result_second.lower()):
                        points += 3

            else:
                player_first = sheet_obj.cell(row=23, column=2 + ((i - 4) * 5)).value
                player_second = sheet_obj.cell(row=23, column=4 + ((i - 4) * 5)).value
                result_first = resultados_sheet.cell(row=23, column=2 + ((i - 4) * 5)).value
                result_second = resultados_sheet.cell(row=23, column=4 + ((i - 4) * 5)).value

                if (result_first != None and result_second != None):

                    if (player_first.lower() == result_first.lower()):
                        points += 3

                    if (player_second.lower() == result_second.lower()):
                        points += 3
        global players_list

        players_list[index]['points_fg'] += points
        players_list[index]['points_total'] += points
        #print(players_list[index]['player name'] + str(players_list[index]['points']))
        index += 1

def octavos():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active
        for i in range(8):

            if (i < 4):
                score1 = sheet_obj.cell(row=27, column=3 + (i * 5)).value
                score2 = sheet_obj.cell(row=27, column=4 + (i * 5)).value
                result_score1 = resultados_sheet.cell(row=27, column=3 + (i * 5)).value
                result_score2 = resultados_sheet.cell(row=27, column=4 + (i * 5)).value
                penalties_1 = resultados_sheet.cell(row=30, column=3 + (i * 5)).value
                penalties_2 = resultados_sheet.cell(row=30, column=4 + (i * 5)).value

                if (result_score1 != None and result_score2 != None and score1 != None and score2 != None):
                    if (score1 != score2):
                        if (score1 == result_score1):
                            points += 2
                        if (score2 == result_score2):
                            points += 2
                    if (score1 > score2 and result_score1 > result_score2):
                        points += 5
                    if (score1 < score2 and result_score1 < result_score2):
                        points += 5
                    if (result_score1 == result_score2):
                        if (score1 > score2 and penalties_1 > penalties_2):
                            points += 5
                        if (score2 > score1 and penalties_2 > penalties_1):
                            points += 5

            else:
                score1 = sheet_obj.cell(row=31, column=3 + (5 *(i - 4))).value
                score2 = sheet_obj.cell(row=31, column=4 + (5 *(i - 4))).value
                result_score1 = resultados_sheet.cell(row=31, column=3 + (5 *(i - 4))).value
                result_score2 = resultados_sheet.cell(row=31, column=4 + (5 *(i - 4))).value
                penalties_1 = resultados_sheet.cell(row=34, column=3 + (5 *(i - 4))).value
                penalties_2 = resultados_sheet.cell(row=34, column=4 + (5 *(i - 4))).value

                if (result_score1 != None and result_score2 != None and score1 != None and score2 != None):
                    if (score1 != score2):
                        if (score1 == result_score1):
                            points += 2
                        if (score2 == result_score2):
                            points += 2
                    if (score1 > score2 and result_score1 > result_score2):
                        points += 5
                    if (score1 < score2 and result_score1 < result_score2):
                        points += 5
                    if (result_score1 == result_score2):
                        if (score1 > score2 and penalties_1 > penalties_2):
                            points += 5
                        if (score2 > score1 and penalties_2 > penalties_1):
                            points += 5

        global players_list

        players_list[index]['points_octavos'] += points
        players_list[index]['points_total'] += points
        index += 1

def cuartos():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active
        for i in range(4):
            score1 = sheet_obj.cell(row=36, column=3 + (i * 5)).value
            score2 = sheet_obj.cell(row=36, column=4 + (i * 5)).value
            result_score1 = resultados_sheet.cell(row=36, column=3 + (i * 5)).value
            result_score2 = resultados_sheet.cell(row=36, column=4 + (i * 5)).value
            penalties_1 = resultados_sheet.cell(row=39, column=3 + (i * 5)).value
            penalties_2 = resultados_sheet.cell(row=39, column=4 + (i * 5)).value

            if(result_score1 != None and result_score2 != None and score1 != None and score2 != None):
                if (score1 == result_score1):
                    points += 2
                if (score2 == result_score2):
                    points += 2
                if (score1 > score2 and result_score1 > result_score2):
                    points += 5
                if (score1 < score2 and result_score1 < result_score2):
                    points += 5
                if (result_score1 == result_score2):
                    if (score1 > score2 and penalties_1 > penalties_2):
                        points += 5
                    if (score2 > score1 and penalties_2 > penalties_1):
                        points += 5

        global players_list

        players_list[index]['points_cuartos'] += points
        players_list[index]['points_total'] += points
        index += 1

def semis():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active
        for i in range(2):
            score1 = sheet_obj.cell(row=41, column=8 + (i * 5)).value
            score2 = sheet_obj.cell(row=41, column=9 + (i * 5)).value
            result_score1 = resultados_sheet.cell(row=41, column=8 + (i * 5)).value
            result_score2 = resultados_sheet.cell(row=41, column=9 + (i * 5)).value
            penalties_1 = resultados_sheet.cell(row=44, column=3 + (i * 5)).value
            penalties_2 = resultados_sheet.cell(row=44, column=4 + (i * 5)).value

            if(result_score1 != None and result_score2 != None and score1 != None and score2 != None):
                if (score1 == result_score1):
                    points += 4
                if (score2 == result_score2):
                    points += 4
                if (score1 > score2 and result_score1 > result_score2):
                    points += 8
                if (score1 < score2 and result_score1 < result_score2):
                    points += 8
                if (result_score1 == result_score2):
                    if (score1 > score2 and penalties_1 > penalties_2):
                        points += 8
                    if (score2 > score1 and penalties_2 > penalties_1):
                        points += 8
        global players_list

        players_list[index]['points_semis'] += points
        players_list[index]['points_total'] += points
        index += 1

def tercerLugar():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active

        score1 = sheet_obj.cell(row=46, column=13).value
        score2 = sheet_obj.cell(row=46, column=14).value
        result_score1 = resultados_sheet.cell(row=46, column=13).value
        result_score2 = resultados_sheet.cell(row=46, column=14).value
        penalties_1 = resultados_sheet.cell(row=49, column=13).value
        penalties_2 = resultados_sheet.cell(row=49, column=14).value

        if (result_score1 != None and result_score2 != None and score1 != None and score2 != None):
            if (score1 == result_score1):
                    points += 4
            if (score2 == result_score2):
                    points += 4
            if (score1 > score2 and result_score1 > result_score2):
                    points += 8
            if (score1 < score2 and result_score1 < result_score2):
                    points += 8
            if (result_score1 == result_score2):
                if (score1 > score2 and penalties_1 > penalties_2):
                    points += 8
                if (score2 > score1 and penalties_2 > penalties_1):
                    points += 8

        global players_list

        players_list[index]['points_tercero'] += points
        players_list[index]['points_total'] += points
        index += 1

def final():
    index = 0
    for path in os.listdir(dir_quinielas):
        points = 0
        wb_obj = openpyxl.load_workbook("quinielas/" + path)
        sheet_obj = wb_obj.active

        score1 = sheet_obj.cell(row=46, column=8).value
        score2 = sheet_obj.cell(row=46, column=9).value
        result_score1 = resultados_sheet.cell(row=46, column=8).value
        result_score2 = resultados_sheet.cell(row=46, column=9).value
        penalties_1 = resultados_sheet.cell(row=49, column=8).value
        penalties_2 = resultados_sheet.cell(row=49, column=9).value

        if (result_score1 != None and result_score2 != None and score1 != None and score2 != None):
            if (score1 == result_score1):
                points += 6
            if (score2 == result_score2):
                points += 6
            if (score1 > score2 and result_score1 > result_score2):
                points += 12
            if (score1 < score2 and result_score1 < result_score2):
                points += 12
            if (result_score1 == result_score2):
                if (score1 > score2 and penalties_1 > penalties_2):
                    points += 12
                if (score2 > score1 and penalties_2 > penalties_1):
                    points += 12

        global players_list

        players_list[index]['points_final'] += points
        players_list[index]['points_total'] += points
        index += 1

def formatoDePosiciones():
    fill_pattern = PatternFill(patternType='solid', fgColor="00008B")
    posiciones_sheet.column_dimensions['A'].width = 30
    posiciones_sheet.column_dimensions['B'].width = 7
    posiciones_sheet.column_dimensions['C'].width = 7
    posiciones_sheet.column_dimensions['D'].width = 7
    posiciones_sheet.column_dimensions['E'].width = 7
    posiciones_sheet.column_dimensions['F'].width = 7
    posiciones_sheet.column_dimensions['G'].width = 7
    posiciones_sheet.column_dimensions['H'].width = 7
    posiciones_sheet['A1'] = "Jugador"
    posiciones_sheet['A1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['A1'].fill = fill_pattern
    posiciones_sheet['A1'].font = Font(color='FFFFFF')
    posiciones_sheet['B1'] = "Grupos"
    posiciones_sheet['B1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['B1'].fill = fill_pattern
    posiciones_sheet['B1'].font = Font(color='FFFFFF')
    posiciones_sheet['C1'] = "Octavos"
    posiciones_sheet['C1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['C1'].fill = fill_pattern
    posiciones_sheet['C1'].font = Font(color='FFFFFF')
    posiciones_sheet['D1'] = "Cuartos"
    posiciones_sheet['D1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['D1'].fill = fill_pattern
    posiciones_sheet['D1'].font = Font(color='FFFFFF')
    posiciones_sheet['E1'] = "Semis"
    posiciones_sheet['E1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['E1'].fill = fill_pattern
    posiciones_sheet['E1'].font = Font(color='FFFFFF')
    posiciones_sheet['F1'] = "3er L"
    posiciones_sheet['F1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['F1'].fill = fill_pattern
    posiciones_sheet['F1'].font = Font(color='FFFFFF')
    posiciones_sheet['G1'] = "Final"
    posiciones_sheet['G1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['G1'].fill = fill_pattern
    posiciones_sheet['G1'].font = Font(color='FFFFFF')
    posiciones_sheet['H1'] = "Total"
    posiciones_sheet['H1'].alignment = Alignment(horizontal="center")
    posiciones_sheet['H1'].fill = fill_pattern
    posiciones_sheet['H1'].font = Font(color='FFFFFF')

def ordenarPorPuntos(players_list):

    return sorted(players_list, key=operator.itemgetter('points_total'), reverse=True)

def poblarPosiciones():
    color_index = 0
    index = 0
    for player_info in players_list:
        if color_index % 2 == 0:
            posiciones_sheet.cell(row=2 + index, column=1).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=2).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=3).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=4).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=5).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=6).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=7).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
            posiciones_sheet.cell(row=2 + index, column=8).fill = PatternFill(start_color='E3E8F0', end_color='E3E8F0',
                                                                              fill_type="solid")
        posiciones_sheet.cell(row=2 + index, column=1).value = player_info['player name']
        posiciones_sheet.cell(row=2 + index, column=1).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=2).value = player_info['points_fg']
        posiciones_sheet.cell(row=2 + index, column=2).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=3).value = player_info['points_octavos']
        posiciones_sheet.cell(row=2 + index, column=3).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=4).value = player_info['points_cuartos']
        posiciones_sheet.cell(row=2 + index, column=4).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=5).value = player_info['points_semis']
        posiciones_sheet.cell(row=2 + index, column=5).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=6).value = player_info['points_tercero']
        posiciones_sheet.cell(row=2 + index, column=6).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=7).value = player_info['points_final']
        posiciones_sheet.cell(row=2 + index, column=7).alignment = Alignment(horizontal='center')
        posiciones_sheet.cell(row=2 + index, column=8).value = player_info['points_total']
        posiciones_sheet.cell(row=2 + index, column=8).alignment = Alignment(horizontal='center')

        color_index += 1
        index += 1


formatoDePosiciones()
faseDeGrupos()
clasificadosOctavos()
octavos()
cuartos()
semis()
tercerLugar()
final()

players_list = ordenarPorPuntos(players_list)
poblarPosiciones()

posiciones.save("posiciones.xlsx")

